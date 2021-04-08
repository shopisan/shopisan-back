from openpyxl import load_workbook
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from store.models import StoreCategories, Store, Address
import json


def create_store_from_row(row, activities, cities):
    store = Store()
    store.name = row[0]
    store.website = row[4]
    if row[5] is not None:
        opening_hours = json.loads(row[5])
        if "all" in opening_hours:
            store.appointmentOnly = True
        else:
            store.openingTimes = opening_hours
    else:
        store.appointmentOnly = True
    store.save()

    add_address_to_store(store, row, cities=cities)

    if isinstance(row[8], int):
        store.categories.add(activities[row[8]])
    else:
        activity_ids = str(row[8]).replace('.', ',').split(',')
        for activity_id in activity_ids:
            if "" != activity_id.strip():
                store.categories.add(activities[int(activity_id)])
        store.save()

    return store


def add_address_to_store(store, row, cities):
    city = ""
    if row[2] in cities and str(row[3]) in cities[row[2]]:
        city = cities[row[2]][str(row[3])]
    address = Address()
    address.streetAvenue = row[1]
    address.country = row[2]
    address.city = city
    address.postalCode = row[3]
    address.store = store
    address.save()


class Command(BaseCommand):
    help = 'Parse Excel files to extract store data'

    def handle(self, *args, **options):
        filename = str(settings.BASE_DIR) + '/store/misc/boutiques.xlsx'
        self.stdout.write(self.style.SUCCESS(print(filename)))
        wb = load_workbook(filename=filename)
        cities = self.parse_cities(workbook=wb)
        activities = self.parse_activities(workbook=wb)

        # activities = {}
        # activities_entity = StoreCategories.objects.all()
        # for cat in activities_entity:
        #     activities[cat.pk] = cat

        self.parse_stores(workbook=wb, activities=activities, cities=cities)
        self.stdout.write(self.style.SUCCESS(""))

    def parse_cities(self, workbook):
        ws = workbook['City']
        cities = {}
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[1] not in cities:
                cities[row[1]] = {}
            cities[row[1]][str(row[0])] = row[2]
        self.stdout.write(self.style.SUCCESS(settings.BASE_DIR))
        return cities

    def parse_activities(self, workbook):
        activities = {}
        ws = workbook['ListActivité']
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            store_cat = StoreCategories()
            store_cat.en = row[1]
            store_cat.fr = row[2]
            store_cat.save()
            activities[store_cat.pk] = store_cat

        self.stdout.write(self.style.SUCCESS("Stores categories imported"))
        return activities

    def parse_stores(self, workbook, activities, cities):
        ws = workbook['Boutiques']

        last_store = None
        for index, row in enumerate(ws.iter_rows(min_row=2, max_col=9, max_row=101, values_only=True)):
            if last_store is None or last_store.name != row[0]:
                store = create_store_from_row(row, activities, cities=cities)
                last_store = store
            else:
                add_address_to_store(store=last_store, row=row, cities=cities)

        self.stdout.write(self.style.SUCCESS("Stores imported"))
